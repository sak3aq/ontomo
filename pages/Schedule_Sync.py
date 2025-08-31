import streamlit as st
import openpyxl
import re
from openpyxl.styles import Font
import io



# --- メインの処理をすべてこの関数に格納 ---
def process_schedule_excel(uploaded_file):
    """
    アップロードされたExcelファイルを処理し、
    処理済みのopenpyxlワークブックオブジェクトを返す。
    エラーが発生した場合はNoneを返す。
    """
    try:
        # メモリ上からExcelファイルを読み込む
        wb = openpyxl.load_workbook(uploaded_file)
        
        # 必要なシートをすべて取得する
        st_syutsuen = wb["出演順"]
        st_meibo = wb["名簿"]
        st_shukketsu = wb["出欠確認"]

    except FileNotFoundError:
        st.error(f"エラー: ファイルが見つかりません。")
        return None
    except KeyError as e:
        st.error(f"エラー: 必要なシート '{e.args[0]}' がファイル内に見つかりません。")
        st.warning("ファイルには「出演順」「名簿」「出欠確認」の3つのシートが必要です。")
        return None

    # --- 2. 各シートからのデータ読み込み ---
    st.info("🔄 STEP 1/6: 各シートからデータを読み込んでいます...")
    list_of_lists = [list(row) for row in st_syutsuen.iter_rows(values_only=True)]
    
    circle_members_data = [list(row) for row in st_meibo.iter_rows(values_only=True)]
    Circle_member_list = []
    for member_i in circle_members_data:
        for member_j in member_i:
            if member_j is not None:
                member_j_cleaned = re.sub(r'\s', '', str(member_j))
                Circle_member_list.append(member_j_cleaned)

    Come_list = [list(row) for row in st_shukketsu.iter_rows(values_only=True)]

    # 「出演順」シートからユニークな出演者リストを作成
    list_all_menber = []
    for i in list_of_lists[1:]:
        if i is None: continue
        for j in i:
            if j is None: continue
            j_cleaned = re.sub(r'\s', '', str(j))
            list_all_menber.append(j_cleaned)
    list_all_menber = sorted(list(set([s for s in list_all_menber if s])))

    # --- 3. 新しいシートの準備 ---
    st.info("🔄 STEP 2/6: 結果を出力する 'After' シートを準備しています...")
    st_name = "After"
    if st_name in wb.sheetnames:
        del wb[st_name]
    st2 = wb.create_sheet(title=st_name)

    # --- 4. 元のデータを新しいシートにコピー ---
    for row_data in list_of_lists:
        st2.append(row_data)

    # --- 5. スケジュールマトリクスの作成 ---
    st.info("🔄 STEP 3/6: スケジュール表を作成し、名簿外のメンバーをチェックしています...")
    # ヘッダー（メンバー名）の書き込みと名簿外メンバーのチェック
    for i, member_name in enumerate(list_all_menber):
        cell = st2.cell(row=1, column=27 + i, value=member_name)
        if member_name not in Circle_member_list:
            cell.font = Font(color='FFFF0000')

    # 出演・シフト状況の書き込みとブッキングチェック
    for row_index, row_data in enumerate(list_of_lists):
        if row_data is None: continue
        cleaned_row_values_syutuen = {re.sub(r'\s', '', str(cell)) for cell in row_data[:9] if cell}
        cleaned_row_values_busyo = {re.sub(r'\s', '', str(cell)) for cell in row_data[9:19] if cell}
        cleaned_row_values_gakki = {re.sub(r'\s', '', str(cell)) for cell in row_data[19:] if cell}
        
        for col_index, member_name in enumerate(list_all_menber):
            target_cell = st2.cell(row=row_index + 1, column=27 + col_index)
            is_booked = False
            if member_name in cleaned_row_values_syutuen:
                target_cell.value = "出演"
                is_booked = True
            
            if member_name in cleaned_row_values_busyo:
                if is_booked:
                    target_cell.value = "ブッキング"
                    target_cell.font = Font(color='FFFF0000')
                else:
                    target_cell.value = "部署シフト"
                    is_booked = True

            if member_name in cleaned_row_values_gakki:
                if is_booked:
                    target_cell.value = "ブッキング"
                    target_cell.font = Font(color='FFFF0000')
                else:
                    target_cell.value = "楽器シフト"

    # --- 6. 出欠確認情報の統合と「欠席ブッキング」のチェック ---
    st.info("🔄 STEP 4/6: 出欠情報を統合しています...")
    after_sheet_data = [list(row) for row in st2.iter_rows(values_only=True)]
    after_header = after_sheet_data[0] if after_sheet_data else []
    
    next_new_col = len(after_header) + 1

    st.info("🔄 STEP 5/6: 欠席ブッキングをチェックしています...")
    for name_index, name_data in enumerate(Come_list[0]):
        if name_data in after_header:
            position = after_header.index(name_data)
            for i in range(1, len(Come_list)):
                cell = st2.cell(row=i + 1, column=position + 1)
                if cell.value is not None:
                    # '×' or '欠席' など、欠席を示す文字列を in でチェック
                    if Come_list[i][name_index] is not None and str(Come_list[i][name_index]).strip() in ["×", "欠席"]:
                        cell.value = "欠席ブッキング"
                        cell.font = Font(color='FFFF0000')
                else:
                    st2.cell(row=i + 1, column=position + 1, value=Come_list[i][name_index])
        else:
            for i in range(len(Come_list)):
                cell_value = Come_list[i][name_index]
                st2.cell(row=i + 1, column=next_new_col, value=cell_value)
            next_new_col += 1
            
    st.info("🔄 STEP 6/6: 処理が完了しました。")
    return wb

# --- Streamlit アプリケーションのUI部分 ---

# アプリのタイトルと説明
st.set_page_config(
    page_title="スケジュール自動チェックツール",
    page_icon="",  
    layout="centered",
    initial_sidebar_state="collapsed"
)
st.title("スケジュール自動チェックツール")
st.markdown("""  
**「出演順」「名簿」「出欠確認」**の3つのシートを含むExcelファイルをアップロードしてください。

**主な機能:**
- **ブッキングチェック**: 同じ時間帯に複数の役割が割り当てられていないか確認します。
- **名簿照合**: スケジュールに記載されているが、公式名簿にない人物を赤字で警告します。
- **欠席ブッキングチェック**: 出演・シフト予定者が「欠席」と回答した場合に警告します。
""")


# ファイルアップローダー
uploaded_file = st.file_uploader(
    "Excelファイル (.xlsx) をここにアップロード",
    type=['xlsx']
)

# ファイルがアップロードされたら処理を開始
if uploaded_file is not None:
    # 処理中のスピナー表示
    with st.spinner('ファイルを処理中です...しばらくお待ちください。'):
        processed_wb = process_schedule_excel(uploaded_file)

    # 処理が成功した場合のみダウンロードボタンを表示
    if processed_wb:
        st.success("🎉 処理が完了しました！")
        
        # 結果をダウンロードするために、メモリ上にExcelファイルを保存
        output_buffer = io.BytesIO()
        processed_wb.save(output_buffer)
        output_buffer.seek(0) # ストリームの先頭にポインタを戻す
        
        # 元のファイル名から新しいファイル名を生成
        original_filename = uploaded_file.name.rsplit('.', 1)[0]
        output_filename = f"{original_filename}_After.xlsx"
        
        # ダウンロードボタン
        st.download_button(
            label="📁 結果のExcelファイルをダウンロード",
            data=output_buffer,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )